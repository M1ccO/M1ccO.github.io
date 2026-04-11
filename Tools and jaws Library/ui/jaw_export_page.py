from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PySide6.QtWidgets import (
    QButtonGroup,
    QDialog,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QTabWidget,
    QVBoxLayout,
)

from data.jaw_database import JawDatabase
from services.jaw_service import JawService
from services.export_service import ExportService
from ui.export_page import ExportPage, ImportMappingDialog


class JawImportMappingDialog(ImportMappingDialog):
    """ImportMappingDialog variant pre-configured for Jaw fields."""

    GENERAL_FIELDS = [
        ('jaw_id', 'Jaw ID'),
        ('jaw_type', 'Jaw type'),
        ('spindle_side', 'Spindle side'),
        ('clamping_diameter_text', 'Clamping diameter'),
        ('clamping_length', 'Clamping length'),
        ('used_in_work', 'Used in works:'),
        ('turning_washer', 'Turning ring'),
        ('last_modified', 'Last modified'),
        ('notes', 'Notes'),
        ('stl_path', '3D model path / JSON'),
    ]
    ADDITIONAL_FIELDS: list = []
    MODEL_FIELDS: list = []

    def __init__(self, excel_headers: list[str], parent=None):
        super().__init__(excel_headers, parent)
        self.setWindowTitle('Import Jaw Mapping')

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(10)

        tabs = QTabWidget()
        tabs.setObjectName('importMappingTabs')
        tabs.addTab(self._build_connector_tab(self.GENERAL_FIELDS), 'General')
        root.addWidget(tabs, 1)

        mode_box = QGroupBox('Import Mode')
        mode_box.setObjectName('importModeBox')
        mode_layout = QVBoxLayout(mode_box)
        self.mode_group = QButtonGroup(self)
        self.mode_overwrite = QRadioButton('Overwrite current database')
        self.mode_overwrite.setChecked(True)
        self.mode_new_db = QRadioButton('Create new database file')
        self.mode_group.addButton(self.mode_overwrite)
        self.mode_group.addButton(self.mode_new_db)

        newdb_row = QHBoxLayout()
        self.new_db_path = QLineEdit()
        self.new_db_path.setPlaceholderText('Choose destination .db file...')
        self.new_db_browse = QPushButton('BROWSE')
        self.new_db_browse.setProperty('panelActionButton', True)
        self.new_db_browse.clicked.connect(self._browse_new_db)
        newdb_row.addWidget(self.new_db_path, 1)
        newdb_row.addWidget(self.new_db_browse)

        mode_layout.addWidget(self.mode_overwrite)
        mode_layout.addWidget(self.mode_new_db)
        mode_layout.addLayout(newdb_row)
        root.addWidget(mode_box)

        self.mode_overwrite.toggled.connect(self._update_mode_ui)
        self._update_mode_ui()

        btns = QHBoxLayout()
        btns.addStretch(1)
        cancel = QPushButton('CANCEL')
        cancel.setProperty('panelActionButton', True)
        ok = QPushButton('IMPORT')
        ok.setProperty('panelActionButton', True)
        ok.setProperty('primaryAction', True)
        cancel.clicked.connect(self.reject)
        ok.clicked.connect(self._accept)
        btns.addWidget(cancel)
        btns.addWidget(ok)
        root.addLayout(btns)

    def _accept(self):
        if 'jaw_id' not in self.mapping():
            QMessageBox.warning(self, 'Import mapping', 'Jaw ID mapping is required.')
            return
        if self.mode_new_db.isChecked() and not self.new_db_path.text().strip():
            QMessageBox.warning(self, 'Import mapping', 'Select destination path for the new database.')
            return
        self.accept()


class _JawToolServiceAdapter:
    def __init__(self, jaw_service: JawService):
        self.jaw_service = jaw_service

    @property
    def db(self):
        return self.jaw_service.db

    def list_tools(self):
        # ExportPage expects list_tools(); map to jaws list API.
        rows = self.jaw_service.list_jaws('', 'all', 'All')
        # Keep row coloring support from ExportService by mirroring jaw_type into tool_type.
        for row in rows:
            row['tool_type'] = row.get('jaw_type', '')
        return rows

    def save_tool(self, jaw: dict):
        self.jaw_service.save_jaw(jaw)


class _JawExportServiceAdapter(ExportService):
    GENERAL_FIELDS = [
        ('jaw_id', 'Jaw ID'),
        ('jaw_type', 'Jaw type'),
        ('spindle_side', 'Spindle side'),
        ('clamping_diameter_text', 'Clamping diameter'),
        ('clamping_length', 'Clamping length'),
        ('used_in_work', 'Used in works:'),
        ('turning_washer', 'Turning ring'),
        ('last_modified', 'Last modified'),
        ('notes', 'Notes'),
        ('stl_path', '3D model path / JSON'),
    ]
    EXPORT_BASE_FIELDS = [
        ('jaw_id', 'Jaw ID'),
        ('jaw_type', 'Jaw type'),
        ('spindle_side', 'Spindle side'),
        ('clamping_diameter_text', 'Clamping diameter'),
        ('clamping_length', 'Clamping length'),
        ('used_in_work', 'Used in works:'),
        ('turning_washer', 'Turning ring'),
        ('last_modified', 'Last modified'),
        ('notes', 'Notes'),
    ]
    EXPORT_HEADER_I18N_KEYS = {
        'jaw_id': 'jaw_library.row.jaw_id',
        'jaw_type': 'jaw_library.row.jaw_type',
        'spindle_side': 'jaw_library.field.spindle_side',
        'clamping_diameter_text': 'jaw_library.row.clamping_diameter',
        'clamping_length': 'jaw_library.row.clamping_length',
        'used_in_work': 'jaw_library.field.used_in_works',
        'turning_washer': 'jaw_library.field.turning_ring',
        'last_modified': 'jaw_library.field.last_modified',
        'notes': 'jaw_library.field.notes',
    }
    _COLUMN_MIN_WIDTHS = {
        'jaw_id': 16,
        'jaw_type': 20,
        'spindle_side': 16,
        'clamping_diameter_text': 22,
        'clamping_length': 18,
        'used_in_work': 24,
        'turning_washer': 16,
        'last_modified': 18,
        'notes': 26,
    }

    IMPORT_DEFAULTS = {
        'jaw_id': '',
        'jaw_type': 'Soft jaws',
        'spindle_side': 'Main spindle',
        'clamping_diameter_text': '',
        'clamping_length': '',
        'used_in_work': '',
        'turning_washer': '',
        'last_modified': '',
        'notes': '',
        'stl_path': '',
    }

    def _localized_jaw_type(self, raw_jaw_type: str) -> str:
        raw = str(raw_jaw_type or '').strip()
        if not raw:
            return raw
        key = f"jaw_library.jaw_type.{raw.lower().replace(' ', '_')}"
        return self._t(key, raw)

    def _localized_spindle_side(self, raw_side: str) -> str:
        raw = str(raw_side or '').strip()
        if not raw:
            return raw
        key = f"jaw_library.spindle_side.{raw.lower().replace(' ', '_')}"
        return self._t(key, raw)

    @staticmethod
    def _jaw_type_bucket(raw_jaw_type: str) -> str:
        value = str(raw_jaw_type or '').strip().casefold()
        if not value:
            return 'soft'
        if ('spik' in value) or ('piikki' in value):
            return 'spiked'
        if ('hard' in value) or ('kova' in value):
            return 'hard'
        if ('special' in value) or ('eriko' in value):
            return 'special'
        if ('soft' in value) or ('pehme' in value):
            return 'soft'
        return 'soft'

    @staticmethod
    def _spindle_bucket(raw_spindle_side: str) -> str:
        value = str(raw_spindle_side or '').strip().casefold()
        if not value:
            return 'main'
        if ('sub' in value) or ('vasta' in value):
            return 'sub'
        if ('both' in value) or ('molem' in value):
            return 'both'
        if ('main' in value) or ('pää' in value) or ('paa' in value):
            return 'main'
        return 'main'

    def _jaw_color_key(self, raw_spindle_side: str, raw_jaw_type: str) -> str:
        spindle = self._spindle_bucket(raw_spindle_side)
        jaw_type = self._jaw_type_bucket(raw_jaw_type)
        return f'{spindle}|{jaw_type}'

    @staticmethod
    def _extract_clamping_metric(raw_value) -> float | None:
        text = str(raw_value or '').strip()
        if not text:
            return None
        nums = re.findall(r'[-+]?\d+(?:[.,]\d+)?', text)
        if not nums:
            return None
        parsed = []
        for token in nums:
            try:
                parsed.append(float(token.replace(',', '.')))
            except Exception:
                continue
        if not parsed:
            return None
        # Use the larger bound for ranges such as "95-100".
        return max(parsed)

    @staticmethod
    def _mix_with_white(hex_rgb: str, ratio: float) -> str:
        ratio = max(0.0, min(0.7, float(ratio)))
        rgb = hex_rgb.strip().lstrip('#')
        if len(rgb) != 6:
            return hex_rgb
        r = int(rgb[0:2], 16)
        g = int(rgb[2:4], 16)
        b = int(rgb[4:6], 16)
        r = int(round(r + (255 - r) * ratio))
        g = int(round(g + (255 - g) * ratio))
        b = int(round(b + (255 - b) * ratio))
        return f'{r:02X}{g:02X}{b:02X}'

    def _jaw_base_color(self, spindle_bucket: str, jaw_type_bucket: str) -> str:
        warm = self._TURNING_TOOLTYPE_ROW_COLORS
        green = self._MILLING_TOOLTYPE_ROW_COLORS

        # Same warm/green families as tool export; keep tones moderate.
        main_colors = {
            'soft': warm[min(3, len(warm) - 1)],
            'hard': warm[min(4, len(warm) - 1)],
            'special': warm[min(5, len(warm) - 1)],
            'spiked': warm[min(6, len(warm) - 1)],
        }
        sub_colors = {
            'special': green[min(4, len(green) - 1)],
            'spiked': green[min(5, len(green) - 1)],
            'hard': green[min(6, len(green) - 1)],
            'soft': green[min(7, len(green) - 1)],
        }
        both_colors = {
            'soft': warm[min(2, len(warm) - 1)],
            'hard': warm[min(3, len(warm) - 1)],
            'special': green[min(3, len(green) - 1)],
            'spiked': green[min(4, len(green) - 1)],
        }

        color_by_bucket = {
            'main': main_colors,
            'sub': sub_colors,
            'both': both_colors,
        }
        spindle_map = color_by_bucket.get(spindle_bucket, main_colors)
        return spindle_map.get(jaw_type_bucket, spindle_map.get('soft', self._UNKNOWN_TOOLTYPE_ROW_COLOR))

    def _tool_type_fill_map(self):
        # Built dynamically by export_tools() from spindle + jaw type + clamping metric.
        return getattr(self, '_jaw_dynamic_fill_map', {})

    @staticmethod
    def _lighten_ratio_for_clamping(metric: float | None) -> float:
        # Anchor near current look at 20-30 mm, and lighten progressively for larger diameters.
        if metric is None:
            return 0.22
        ratio = 0.20 + ((float(metric) - 25.0) * 0.0028)
        return max(0.14, min(0.58, ratio))

    def export_tools(self, filename: str, tools: list[dict]):
        rows_in = list(tools or [])

        self._jaw_dynamic_fill_map = {}
        rows = []
        for row in rows_in:
            converted = dict(row)
            raw_jaw_type = row.get('jaw_type', '')
            raw_spindle_side = row.get('spindle_side', '')
            converted['jaw_type'] = self._localized_jaw_type(row.get('jaw_type', ''))
            converted['spindle_side'] = self._localized_spindle_side(row.get('spindle_side', ''))

            spindle_bucket = self._spindle_bucket(raw_spindle_side)
            jaw_type_bucket = self._jaw_type_bucket(raw_jaw_type)
            base_color = self._jaw_base_color(spindle_bucket, jaw_type_bucket)

            metric = self._extract_clamping_metric(row.get('clamping_diameter_text', ''))
            lighten = self._lighten_ratio_for_clamping(metric)

            final_color = self._mix_with_white(base_color, lighten)
            style_key = f'{spindle_bucket}|{jaw_type_bucket}|{final_color}'.casefold()
            self._jaw_dynamic_fill_map[style_key] = PatternFill(fill_type='solid', fgColor=final_color)
            converted['tool_type'] = style_key
            rows.append(converted)
        wb = Workbook()
        ws = wb.active
        ws.title = 'JAWS'
        self._write_tools_sheet(ws, rows)
        wb.save(filename)


class JawExportPage(ExportPage):
    def __init__(self, jaw_service, on_jaw_data_changed=None, on_jaw_database_switched=None, parent=None, translate=None):
        self.jaw_service = jaw_service
        self._jaw_tool_adapter = _JawToolServiceAdapter(jaw_service)
        self._jaw_export_adapter = _JawExportServiceAdapter()
        super().__init__(
            tool_service=self._jaw_tool_adapter,
            export_service=self._jaw_export_adapter,
            on_data_changed=on_jaw_data_changed,
            on_database_switched=on_jaw_database_switched,
            parent=parent,
            translate=translate,
        )

    def set_jaw_service(self, jaw_service: JawService):
        self.jaw_service = jaw_service
        self._jaw_tool_adapter.jaw_service = jaw_service

    def _export_filename_prefix(self) -> str:
        return 'jaws-library-export'

    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, self._t('jaw_library.import.title', 'Import from Excel'), '', self._t('jaw_library.import.filter_excel', 'Excel (*.xlsx *.xlsm)'))
        if not path:
            return

        try:
            headers = self.export_service.read_excel_headers(path)
        except Exception as exc:
            QMessageBox.critical(self, self._t('jaw_library.import.failed_title', 'Import failed'), self._t('jaw_library.import.read_headers_failed', 'Could not read Excel headers:\n{error}', error=exc))
            return

        if not headers:
            QMessageBox.warning(self, 'Import', 'No header row found in selected Excel file.')
            return

        dlg = JawImportMappingDialog(headers, self)
        if dlg.exec() != QDialog.Accepted:
            return

        mapping = dlg.mapping()
        if 'jaw_id' not in mapping:
            QMessageBox.warning(self, 'Import mapping', 'Jaw ID mapping is required.')
            return

        try:
            jaws = self.export_service.import_tools(path, mapping)
        except Exception as exc:
            QMessageBox.critical(self, 'Import failed', str(exc))
            return

        if not jaws:
            QMessageBox.information(self, 'Import', 'No valid jaw rows found with current mapping.')
            return

        try:
            if dlg.import_mode() == 'overwrite':
                confirm = QMessageBox.question(
                    self,
                    'Confirm overwrite',
                    'Are you sure you want to overwrite current database contents?',
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.Yes,
                )
                if confirm != QMessageBox.Yes:
                    return

                backup_answer = QMessageBox.question(
                    self,
                    'Backup before overwrite',
                    'Create a backup of the current database before overwrite?',
                    QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                    QMessageBox.Yes,
                )
                if backup_answer == QMessageBox.Cancel:
                    return
                if backup_answer == QMessageBox.Yes:
                    backup_path = self._create_database_backup()
                    QMessageBox.information(self, 'Backup created', f'Backup saved to:\n{backup_path}')

                with self.tool_service.db.conn:
                    self.tool_service.db.conn.execute('DELETE FROM jaws')
                    for jaw in jaws:
                        self.tool_service.save_tool(jaw)
                if callable(self.on_data_changed):
                    self.on_data_changed()
                QMessageBox.information(self, 'Import', f'Imported {len(jaws)} jaws into current database.')
                return

            db_path = Path(dlg.selected_new_db_path())
            if db_path.suffix.lower() != '.db':
                db_path = db_path.with_suffix('.db')
            new_db = JawDatabase(db_path)
            new_jaw_service = JawService(new_db)
            with new_db.conn:
                new_db.conn.execute('DELETE FROM jaws')
            for jaw in jaws:
                new_jaw_service.save_jaw(jaw)
            new_db.close()
            QMessageBox.information(self, 'Import', f'Imported {len(jaws)} jaws into new database:\n{db_path}')
        except Exception as exc:
            QMessageBox.critical(self, 'Import failed', str(exc))
