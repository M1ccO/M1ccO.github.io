import json

from config import ALL_TOOL_TYPES, I18N_DIR, MILLING_TOOL_TYPES, TURNING_TOOL_TYPES
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExportService:
    # Backward-compatible mapping fields used by import flow.
    GENERAL_FIELDS = [
        ('id', 'Tool ID'),
        ('tool_head', 'Tool Head'),
        ('tool_type', 'Tool type'),
        ('description', 'Description'),
        ('geom_x', 'Geom X'),
        ('geom_z', 'Geom Z'),
        ('radius', 'Radius'),
        ('nose_corner_radius', 'Nose R / Corner R'),
        ('holder_code', 'Holder code'),
        ('holder_link', 'Holder link'),
        ('holder_add_element', 'Add. Element'),
        ('holder_add_element_link', 'Add. Element link'),
        ('cutting_type', 'Cutting component type'),
        ('cutting_code', 'Cutting component code'),
        ('cutting_link', 'Cutting component link'),
        ('cutting_add_element', 'Add. Insert/Drill/Mill'),
        ('cutting_add_element_link', 'Add. Insert/Drill/Mill link'),
        ('drill_nose_angle', 'Nose angle'),
        ('mill_cutting_edges', 'Cutting edges'),
        ('notes', 'Notes'),
    ]

    # Tool export schema (HEAD split by worksheet).
    EXPORT_BASE_FIELDS = [
        ('id', 'Tool ID'),
        ('description', 'Description'),
        ('tool_type', 'Tool type'),
        ('geom_x', 'Geom X'),
        ('geom_z', 'Geom Z'),
        ('radius', 'Radius'),
        ('nose_corner_radius', 'Nose R / Corner R'),
        ('drill_nose_angle', 'Nose angle'),
        ('mill_cutting_edges', 'Cutting edges'),
        ('notes', 'Notes'),
        ('holder_code', 'Holder code'),
        ('cutting_code', 'Cutting code'),
    ]
    EXPORT_HEADER_I18N_KEYS = {
        'id': 'tool_library.export.header.tool_id',
        'description': 'tool_library.export.header.description',
        'tool_type': 'tool_library.export.header.tool_type',
        'geom_x': 'tool_library.export.header.geom_x',
        'geom_z': 'tool_library.export.header.geom_z',
        'radius': 'tool_library.export.header.radius',
        'nose_corner_radius': 'tool_library.export.header.nose_corner_radius',
        'notes': 'tool_library.export.header.notes',
        'holder_code': 'tool_library.export.header.holder_code',
        'cutting_code': 'tool_library.export.header.cutting_code',
        'drill_nose_angle': 'tool_library.export.header.nose_angle',
        'mill_cutting_edges': 'tool_library.export.header.cutting_edges',
    }

    _ROLE_ORDER = ('holder', 'cutting', 'support')
    _ROLE_LABEL = {
        'holder': 'Holder',
        'cutting': 'Cutting',
        'support': 'Support',
    }
    _COMPONENT_SLOT_CAP = 2
    _TURNING_TOOLTYPE_ROW_COLORS = [
        'FFF4E8',
        'FFEFD9',
        'FFE9CB',
        'FFE3BC',
        'FFDEAE',
        'FFD89F',
        'FFD291',
        'FFCD82',
        'FFC774',
        'FFC165',
    ]
    _MILLING_TOOLTYPE_ROW_COLORS = [
        'EAF8F2',
        'E3F5EE',
        'DCF2EA',
        'D5EFE6',
        'CEECE2',
        'C7E9DE',
        'C0E6DA',
        'B9E3D6',
        'B2E0D2',
        'ABDCCE',
        'A4D9CA',
        '9DD6C6',
    ]
    _UNKNOWN_TOOLTYPE_ROW_COLOR = 'ECEFF3'
    _CUTTING_TYPES = ('Insert', 'Drill', 'Mill')
    _DRILL_LIKE_TOOL_TYPES = {'Drill', 'Spot Drill', 'Reamer', 'Tapping', 'Turn Drill', 'Turn Spot Drill'}
    _HEADER_FILL_COLOR = 'CFE4F8'
    _HEADER_FONT_COLOR = '16334E'
    _HEADER_BORDER_COLOR = 'D0D7DE'
    _COLUMN_MIN_WIDTHS = {
        'id': 18,
        'description': 34,
        'tool_type': 24,
        'geom_x': 15,
        'geom_z': 15,
        'radius': 10,
        'nose_corner_radius': 32,
        'drill_nose_angle': 18,
        'mill_cutting_edges': 18,
        'notes': 26,
        'holder_code': 24,
        'cutting_code': 24,
    }

    IMPORT_DEFAULTS = {
        'id': '',
        'tool_head': 'HEAD1',
        'tool_type': 'O.D Turning',
        'description': '',
        'geom_x': 0.0,
        'geom_z': 0.0,
        'radius': 0.0,
        'nose_corner_radius': 0.0,
        'holder_code': '',
        'holder_link': '',
        'holder_add_element': '',
        'holder_add_element_link': '',
        'cutting_type': 'Insert',
        'cutting_code': '',
        'cutting_link': '',
        'cutting_add_element': '',
        'cutting_add_element_link': '',
        'notes': '',
        'drill_nose_angle': 0.0,
        'mill_cutting_edges': 0,
        'geometry_profiles': [],
        'support_parts': [],
        'stl_path': '',
    }

    def __init__(self, translate=None):
        self._translate = translate or (lambda _key, default=None, **_kwargs: default or '')
        self._i18n_cache: dict[str, dict] = {}
        self._tool_type_alias_map: dict[str, str] = {}
        self._cutting_type_alias_map: dict[str, str] = {}
        self._rebuild_import_alias_maps()

    def set_translator(self, translate):
        self._translate = translate or (lambda _key, default=None, **_kwargs: default or '')
        self._rebuild_import_alias_maps()

    def _t(self, key: str, default: str) -> str:
        try:
            return self._translate(key, default)
        except Exception:
            return default

    def _localized_tool_type(self, raw_tool_type: str) -> str:
        raw = str(raw_tool_type or '').strip()
        if not raw:
            return raw
        key = f"tool_library.tool_type.{raw.lower().replace('.', '_').replace('/', '_').replace(' ', '_')}"
        return self._t(key, raw)

    def _localized_cutting_type(self, raw_cutting_type: str) -> str:
        raw = str(raw_cutting_type or '').strip()
        if not raw:
            return raw
        key = f"tool_library.cutting_type.{raw.lower().replace(' ', '_')}"
        return self._t(key, raw)

    @staticmethod
    def _lookup_key(value) -> str:
        return ' '.join(str(value or '').strip().casefold().split())

    @staticmethod
    def _coerce_text(value) -> str:
        if value is None:
            return ''
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value).strip()
        return str(value).strip()

    @staticmethod
    def _coerce_float(value, default: float = 0.0) -> float:
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip()
        if not text:
            return default
        text = text.replace(' ', '').replace(',', '.')
        try:
            return float(text)
        except Exception:
            return default

    @staticmethod
    def _coerce_int(value, default: int = 0) -> int:
        if value is None:
            return default
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        text = str(value).strip()
        if not text:
            return default
        try:
            return int(float(text.replace(',', '.')))
        except Exception:
            return default

    def _load_translation_catalog(self, language: str) -> dict:
        lang = str(language or '').strip().lower()
        if not lang:
            return {}
        if lang in self._i18n_cache:
            return self._i18n_cache[lang]
        path = I18N_DIR / f'{lang}.json'
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if not isinstance(data, dict):
                data = {}
        except Exception:
            data = {}
        self._i18n_cache[lang] = data
        return data

    @staticmethod
    def _tool_type_i18n_key(raw_tool_type: str) -> str:
        normalized = (raw_tool_type or '').strip().lower().replace('.', '_').replace('/', '_').replace(' ', '_')
        return f"tool_library.tool_type.{normalized}"

    @staticmethod
    def _cutting_type_i18n_key(raw_cutting_type: str) -> str:
        normalized = (raw_cutting_type or '').strip().lower().replace(' ', '_')
        return f"tool_library.cutting_type.{normalized}"

    @staticmethod
    def _add_alias(alias_map: dict[str, str], alias_value, canonical: str):
        key = ExportService._lookup_key(alias_value)
        if key:
            alias_map[key] = canonical

    def _rebuild_import_alias_maps(self):
        tool_aliases: dict[str, str] = {}
        cutting_aliases: dict[str, str] = {}

        en_catalog = self._load_translation_catalog('en')
        fi_catalog = self._load_translation_catalog('fi')

        for canonical in ALL_TOOL_TYPES:
            key = self._tool_type_i18n_key(canonical)
            self._add_alias(tool_aliases, canonical, canonical)
            self._add_alias(tool_aliases, self._t(key, canonical), canonical)
            self._add_alias(tool_aliases, en_catalog.get(key, ''), canonical)
            self._add_alias(tool_aliases, fi_catalog.get(key, ''), canonical)

        for canonical in self._CUTTING_TYPES:
            key = self._cutting_type_i18n_key(canonical)
            self._add_alias(cutting_aliases, canonical, canonical)
            self._add_alias(cutting_aliases, self._t(key, canonical), canonical)
            self._add_alias(cutting_aliases, en_catalog.get(key, ''), canonical)
            self._add_alias(cutting_aliases, fi_catalog.get(key, ''), canonical)

        self._tool_type_alias_map = tool_aliases
        self._cutting_type_alias_map = cutting_aliases

    def _normalize_tool_type_value(self, value: str) -> str:
        raw = self._coerce_text(value)
        if not raw:
            return raw
        return self._tool_type_alias_map.get(self._lookup_key(raw), raw)

    def _normalize_cutting_type_value(self, value: str) -> str:
        raw = self._coerce_text(value)
        if not raw:
            return ''
        return self._cutting_type_alias_map.get(self._lookup_key(raw), '')

    @staticmethod
    def _cell_hyperlink_target(cell) -> str:
        if cell is None:
            return ''
        link = getattr(cell, 'hyperlink', None)
        if not link:
            return ''
        target = getattr(link, 'target', None)
        return str(target or '').strip()

    @staticmethod
    def _mapped_cell(row_cells, header_to_idx: dict[str, int], mapping: dict[str, str], field_key: str):
        excel_header = str(mapping.get(field_key, '') or '').strip()
        if not excel_header:
            return None
        idx = header_to_idx.get(excel_header)
        if idx is None or idx < 0 or idx >= len(row_cells):
            return None
        return row_cells[idx]

    def _infer_cutting_type(self, tool: dict) -> str:
        if self._coerce_int(tool.get('mill_cutting_edges', 0), 0) > 0:
            return 'Mill'
        if self._coerce_float(tool.get('drill_nose_angle', 0), 0.0) > 0:
            return 'Drill'

        normalized_tool_type = self._normalize_tool_type_value(tool.get('tool_type', ''))
        if normalized_tool_type in self._DRILL_LIKE_TOOL_TYPES:
            return 'Drill'
        if normalized_tool_type in MILLING_TOOL_TYPES:
            return 'Mill'
        return 'Insert'

    def _component_items_from_import_row(self, tool: dict) -> list[dict]:
        holder_code = self._coerce_text(tool.get('holder_code', ''))
        holder_link = self._coerce_text(tool.get('holder_link', ''))
        cutting_code = self._coerce_text(tool.get('cutting_code', ''))
        cutting_link = self._coerce_text(tool.get('cutting_link', ''))
        cutting_type = self._coerce_text(tool.get('cutting_type', 'Insert')) or 'Insert'

        items = []
        if holder_code:
            items.append(
                {
                    'role': 'holder',
                    'label': self._t('tool_library.field.holder', 'Holder'),
                    'code': holder_code,
                    'link': holder_link,
                    'group': '',
                    'order': len(items),
                }
            )

        if cutting_code:
            items.append(
                {
                    'role': 'cutting',
                    'label': self._localized_cutting_type(cutting_type),
                    'code': cutting_code,
                    'link': cutting_link,
                    'group': '',
                    'order': len(items),
                }
            )
        return items

    def _normalize_number(self, value):
        if value is None or value == '':
            return None
        try:
            return float(value)
        except Exception:
            return value

    @staticmethod
    def _normalize_tool_head(value) -> str:
        head = str(value or 'HEAD1').strip().upper()
        return head if head in {'HEAD1', 'HEAD2'} else 'HEAD1'

    @staticmethod
    def _parse_json_list_or_empty(value):
        if isinstance(value, list):
            return value
        if value is None:
            return []
        text = str(value).strip()
        if not text:
            return []
        import json
        try:
            parsed = json.loads(text)
            return parsed if isinstance(parsed, list) else []
        except Exception:
            return []

    def _normalize_component_items(self, tool: dict) -> list[dict]:
        items = self._parse_json_list_or_empty(tool.get('component_items', []))

        normalized = []
        for idx, item in enumerate(items):
            if not isinstance(item, dict):
                continue
            role = str(item.get('role') or '').strip().lower()
            code = str(item.get('code') or '').strip()
            if role not in self._ROLE_ORDER or not code:
                continue
            label = str(item.get('label') or '').strip() or self._ROLE_LABEL[role]
            try:
                order = int(item.get('order', idx))
            except Exception:
                order = idx
            normalized.append(
                {
                    'role': role,
                    'label': label,
                    'code': code,
                    'link': str(item.get('link') or '').strip(),
                    'group': str(item.get('group') or '').strip(),
                    'order': order,
                }
            )

        if normalized:
            normalized.sort(key=lambda entry: int(entry.get('order', 0)))
            return normalized

        fallback = []

        def _add(role, label, code, link='', group=''):
            code_text = str(code or '').strip()
            if not code_text:
                return
            fallback.append(
                {
                    'role': role,
                    'label': str(label or '').strip() or self._ROLE_LABEL[role],
                    'code': code_text,
                    'link': str(link or '').strip(),
                    'group': str(group or '').strip(),
                    'order': len(fallback),
                }
            )

        cutting_type = str(tool.get('cutting_type', 'Insert') or 'Insert').strip() or 'Insert'
        _add('holder', 'Holder', tool.get('holder_code', ''), tool.get('holder_link', ''))
        _add('holder', 'Add. Element', tool.get('holder_add_element', ''), tool.get('holder_add_element_link', ''))
        _add('cutting', cutting_type, tool.get('cutting_code', ''), tool.get('cutting_link', ''))
        _add(
            'cutting',
            f'Add. {cutting_type}',
            tool.get('cutting_add_element', ''),
            tool.get('cutting_add_element_link', ''),
        )

        for part in self._parse_json_list_or_empty(tool.get('support_parts', [])):
            if isinstance(part, str):
                try:
                    part = json.loads(part)
                except Exception:
                    part = {'name': part, 'code': '', 'link': '', 'group': ''}
            if not isinstance(part, dict):
                continue
            _add(
                'support',
                str(part.get('name') or 'Part').strip() or 'Part',
                part.get('code', ''),
                part.get('link', ''),
                part.get('group', ''),
            )

        return fallback

    def _component_export_payload(self, tool: dict) -> dict:
        items = self._normalize_component_items(tool)

        holder_idx = next((idx for idx, item in enumerate(items) if item.get('role') == 'holder'), None)
        cutting_idx = next((idx for idx, item in enumerate(items) if item.get('role') == 'cutting'), None)

        holder = items[holder_idx] if holder_idx is not None else None
        cutting = items[cutting_idx] if cutting_idx is not None else None

        def _text(item: dict | None, key: str) -> str:
            if not item:
                return ''
            return str(item.get(key) or '').strip()

        return {
            'holder_code': _text(holder, 'code'),
            'holder_link': _text(holder, 'link'),
            'cutting_code': _text(cutting, 'code'),
            'cutting_link': _text(cutting, 'link'),
        }

    def _analyze_component_layout(self, tools: list[dict], slot_cap: int | None = None):
        cap = int(slot_cap or self._COMPONENT_SLOT_CAP)
        max_counts = {role: 0 for role in self._ROLE_ORDER}
        overflow = {role: False for role in self._ROLE_ORDER}

        for tool in tools or []:
            items = self._normalize_component_items(tool)
            by_role = {role: [] for role in self._ROLE_ORDER}
            for item in items:
                role = item.get('role')
                if role in by_role:
                    by_role[role].append(item)
            for role in self._ROLE_ORDER:
                count = len(by_role[role])
                max_counts[role] = max(max_counts[role], count)
                if count > cap:
                    overflow[role] = True

        slot_counts = {role: min(cap, max_counts[role]) for role in self._ROLE_ORDER}
        return {
            'slot_counts': slot_counts,
            'overflow': overflow,
            'slot_cap': cap,
        }

    def _component_columns(self, layout):
        columns = []
        for role in self._ROLE_ORDER:
            role_label = self._ROLE_LABEL[role]
            for slot in range(1, int(layout['slot_counts'][role]) + 1):
                columns.extend(
                    [
                        {
                            'key': f'{role}_{slot}_label',
                            'label': f'{role_label} {slot} label',
                            'role': role,
                            'slot': slot,
                            'attr': 'label',
                        },
                        {
                            'key': f'{role}_{slot}_code',
                            'label': f'{role_label} {slot} code',
                            'role': role,
                            'slot': slot,
                            'attr': 'code',
                        },
                        {
                            'key': f'{role}_{slot}_link',
                            'label': f'{role_label} {slot} link',
                            'role': role,
                            'slot': slot,
                            'attr': 'link',
                        },
                        {
                            'key': f'{role}_{slot}_group',
                            'label': f'{role_label} {slot} group',
                            'role': role,
                            'slot': slot,
                            'attr': 'group',
                        },
                    ]
                )

            if layout['overflow'][role]:
                columns.append(
                    {
                        'key': f'{role}_overflow_json',
                        'label': f'{role_label} overflow (JSON)',
                        'role': role,
                        'slot': None,
                        'attr': 'overflow',
                    }
                )
        return columns

    @staticmethod
    def _group_items_by_role(items: list[dict]):
        grouped = {'holder': [], 'cutting': [], 'support': []}
        for item in items or []:
            role = item.get('role')
            if role in grouped:
                grouped[role].append(item)
        for role in grouped:
            grouped[role].sort(key=lambda entry: int(entry.get('order', 0)))
        return grouped

    def _tool_type_fill_map(self) -> dict[str, PatternFill]:
        fills: dict[str, PatternFill] = {}

        for idx, tool_type in enumerate(TURNING_TOOL_TYPES):
            color = self._TURNING_TOOLTYPE_ROW_COLORS[idx % len(self._TURNING_TOOLTYPE_ROW_COLORS)]
            fills[str(tool_type).strip().casefold()] = PatternFill(fill_type='solid', fgColor=color)

        for idx, tool_type in enumerate(MILLING_TOOL_TYPES):
            color = self._MILLING_TOOLTYPE_ROW_COLORS[idx % len(self._MILLING_TOOLTYPE_ROW_COLORS)]
            fills[str(tool_type).strip().casefold()] = PatternFill(fill_type='solid', fgColor=color)

        return fills

    def read_excel_headers(self, filename: str) -> list[str]:
        wb = load_workbook(filename=filename, read_only=True, data_only=True)
        try:
            ws = wb.active
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not first_row:
                return []
            return [str(v).strip() for v in first_row if v is not None and str(v).strip()]
        finally:
            wb.close()

    def import_tools(self, filename: str, mapping: dict[str, str]) -> list[dict]:
        wb = load_workbook(filename=filename, read_only=False, data_only=True)
        try:
            ws = wb.active
            raw_headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not raw_headers:
                return []
            headers = [str(v).strip() if v is not None else '' for v in raw_headers]
            header_to_idx = {h: i for i, h in enumerate(headers) if h}

            float_fields = {'geom_x', 'geom_z', 'radius', 'nose_corner_radius', 'drill_nose_angle'}
            int_fields = {'mill_cutting_edges'}
            list_fields = {'support_parts', 'geometry_profiles'}
            required_id_field = 'id' if 'id' in self.IMPORT_DEFAULTS else ('jaw_id' if 'jaw_id' in self.IMPORT_DEFAULTS else '')
            is_tool_import = all(
                field in self.IMPORT_DEFAULTS
                for field in ('tool_type', 'cutting_type', 'holder_code', 'cutting_code')
            )

            imported = []
            for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=False):
                tool = dict(self.IMPORT_DEFAULTS)

                for field_key, excel_header in mapping.items():
                    if field_key not in self.IMPORT_DEFAULTS:
                        continue
                    idx = header_to_idx.get(str(excel_header or '').strip())
                    if idx is None or idx >= len(row):
                        continue
                    cell = row[idx]
                    raw = cell.value if cell is not None else None
                    if raw is None:
                        continue

                    if field_key in float_fields:
                        tool[field_key] = self._coerce_float(raw, float(tool.get(field_key, 0.0) or 0.0))
                    elif field_key in int_fields:
                        tool[field_key] = self._coerce_int(raw, int(tool.get(field_key, 0) or 0))
                    elif field_key in list_fields:
                        tool[field_key] = self._parse_json_list_or_empty(raw)
                    else:
                        tool[field_key] = self._coerce_text(raw)

                if 'tool_head' in self.IMPORT_DEFAULTS:
                    tool['tool_head'] = self._normalize_tool_head(tool.get('tool_head', 'HEAD1'))

                if is_tool_import:
                    holder_code_cell = self._mapped_cell(row, header_to_idx, mapping, 'holder_code')
                    cutting_code_cell = self._mapped_cell(row, header_to_idx, mapping, 'cutting_code')
                    cutting_type_cell = self._mapped_cell(row, header_to_idx, mapping, 'cutting_type')
                    holder_hyperlink = self._cell_hyperlink_target(holder_code_cell)
                    cutting_hyperlink = self._cell_hyperlink_target(cutting_code_cell)

                    holder_link_column = self._coerce_text(tool.get('holder_link', ''))
                    cutting_link_column = self._coerce_text(tool.get('cutting_link', ''))
                    tool['holder_link'] = holder_link_column or holder_hyperlink
                    tool['cutting_link'] = cutting_link_column or cutting_hyperlink

                    tool['tool_type'] = self._normalize_tool_type_value(tool.get('tool_type', ''))

                    mapped_cutting_raw = self._coerce_text(cutting_type_cell.value if cutting_type_cell is not None else '')
                    explicit_cutting_type = self._normalize_cutting_type_value(mapped_cutting_raw)
                    tool['cutting_type'] = explicit_cutting_type or self._infer_cutting_type(tool)
                    tool['component_items'] = self._component_items_from_import_row(tool)

                if required_id_field and not self._coerce_text(tool.get(required_id_field, '')):
                    continue

                imported.append(tool)

            return imported
        finally:
            wb.close()

    def _write_tools_sheet(self, ws, tools: list[dict]):
        headers = [
            self._t(self.EXPORT_HEADER_I18N_KEYS.get(key, key), label)
            for key, label in self.EXPORT_BASE_FIELDS
        ]

        # Header row
        ws.append(headers)
        header_fill = PatternFill(fill_type='solid', fgColor=self._HEADER_FILL_COLOR)
        header_font = Font(name='Segoe UI', color=self._HEADER_FONT_COLOR, bold=True)
        thin = Side(style='thin', color='D0D7DE')
        header_side = Side(style='thin', color=self._HEADER_BORDER_COLOR)
        header_border = Border(left=header_side, right=header_side, top=header_side, bottom=header_side)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        default_font = Font(name='Segoe UI')
        hyperlink_font = Font(name='Segoe UI', color='0563C1', underline='single')

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = header_border

        tool_type_fill = self._tool_type_fill_map()
        unknown_fill = PatternFill(fill_type='solid', fgColor=self._UNKNOWN_TOOLTYPE_ROW_COLOR)

        numeric_keys = {'geom_x', 'geom_z', 'radius', 'nose_corner_radius', 'drill_nose_angle'}
        int_keys = {'mill_cutting_edges'}
        hyperlink_targets = {
            'holder_code': 'holder_link',
            'cutting_code': 'cutting_link',
        }

        # Data rows
        for tool in tools or []:
            normalized_tool = dict(tool)
            normalized_tool.update(self._component_export_payload(normalized_tool))

            row_values = []
            for key, _label in self.EXPORT_BASE_FIELDS:
                value = normalized_tool.get(key, '')
                if key == 'tool_type':
                    value = self._localized_tool_type(value)
                if key in numeric_keys:
                    value = self._normalize_number(value)
                elif key in int_keys:
                    try:
                        value = int(value or 0)
                    except Exception:
                        value = value
                row_values.append(value)

            ws.append(row_values)
            row_idx = ws.max_row

            type_key = (tool.get('tool_type', '') or '').strip().casefold()
            fill = tool_type_fill.get(type_key, unknown_fill)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.fill = fill

                key = self.EXPORT_BASE_FIELDS[col_idx - 1][0]
                if key in numeric_keys and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.000'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif key in int_keys and isinstance(cell.value, int):
                    cell.number_format = '0'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = default_font

                link_key = hyperlink_targets.get(key)
                if link_key:
                    link_value = str(normalized_tool.get(link_key, '') or '').strip()
                    if link_value:
                        try:
                            cell.hyperlink = link_value
                            cell.font = hyperlink_font
                        except Exception:
                            pass

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        # Auto-size columns with sensible min/max limits.
        for col_idx in range(1, len(headers) + 1):
            letter = get_column_letter(col_idx)
            max_len = len(headers[col_idx - 1])
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            key = self.EXPORT_BASE_FIELDS[col_idx - 1][0]
            min_width = int(self._COLUMN_MIN_WIDTHS.get(key, 10))
            width = max(max_len + 6, min_width)
            ws.column_dimensions[letter].width = min(width, 64)

        # Slightly taller rows for readability.
        ws.row_dimensions[1].height = 24
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 21

    def export_tools(self, filename: str, tools: list[dict]):
        wb = Workbook()
        head1_ws = wb.active
        head1_ws.title = 'HEAD1'
        head2_ws = wb.create_sheet('HEAD2')

        head1_tools = []
        head2_tools = []

        for tool in tools or []:
            normalized_head = self._normalize_tool_head(tool.get('tool_head', 'HEAD1'))
            tool_copy = dict(tool)
            tool_copy['tool_head'] = normalized_head
            if normalized_head == 'HEAD2':
                head2_tools.append(tool_copy)
            else:
                head1_tools.append(tool_copy)

        self._write_tools_sheet(head1_ws, head1_tools)
        self._write_tools_sheet(head2_ws, head2_tools)

        wb.save(filename)

    def export_tools_to_excel(self, conn, filename):
        # Backward-compatible path for legacy callers.
        rows = conn.execute('SELECT * FROM tools ORDER BY id').fetchall()
        tools = [dict(r) for r in rows]
        self.export_tools(filename, tools)
