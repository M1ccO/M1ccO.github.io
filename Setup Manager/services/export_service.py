from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExportService:
    # Export only fields shown in Tool Editor -> General tab.
    GENERAL_FIELDS = [
        ('id', 'Tool ID'),
        ('tool_type', 'Tool type'),
        ('description', 'Description'),
        ('geom_x', 'Geom X'),
        ('geom_z', 'Geom Z'),
        ('radius', 'Radius'),
        ('nose_corner_radius', 'Nose R / Corner R'),
        ('holder_code', 'Holder code'),
        ('holder_link', 'Holder link'),
        ('holder_add_element', 'Add. Element'),
        ('holder_add_element_link', 'Add. Element link'),
        ('cutting_type', 'Cutting component type'),
        ('cutting_code', 'Cutting component code'),
        ('cutting_link', 'Cutting component link'),
        ('cutting_add_element', 'Add. Insert/Drill/Mill'),
        ('cutting_add_element_link', 'Add. Insert/Drill/Mill link'),
        ('drill_nose_angle', 'Nose angle'),
        ('mill_cutting_edges', 'Cutting edges'),
        ('notes', 'Notes'),
    ]

    _TOOLTYPE_ROW_COLORS = [
        'EAF4FF',  # light blue
        'EDF8EA',  # light green
        'FFF4E8',  # light orange
        'F5ECFF',  # light violet
        'EAF8F8',  # light cyan
        'FFF9E3',  # light yellow
        'FDEEEF',  # light rose
        'ECEFF3',  # light slate
    ]

    IMPORT_DEFAULTS = {
        'id': '',
        'tool_type': 'O.D Turning',
        'description': '',
        'geom_x': 0.0,
        'geom_z': 0.0,
        'radius': 0.0,
        'nose_corner_radius': 0.0,
        'holder_code': '',
        'holder_link': '',
        'holder_add_element': '',
        'holder_add_element_link': '',
        'cutting_type': 'Insert',
        'cutting_code': '',
        'cutting_link': '',
        'cutting_add_element': '',
        'cutting_add_element_link': '',
        'notes': '',
        'drill_nose_angle': 0.0,
        'mill_cutting_edges': 0,
        'geometry_profiles': [],
        'support_parts': [],
        'stl_path': '',
    }

    def _normalize_number(self, value):
        if value is None or value == '':
            return None
        try:
            return float(value)
        except Exception:
            return value

    @staticmethod
    def _parse_json_list_or_empty(value):
        if isinstance(value, list):
            return value
        if value is None:
            return []
        text = str(value).strip()
        if not text:
            return []
        import json
        try:
            parsed = json.loads(text)
            return parsed if isinstance(parsed, list) else []
        except Exception:
            return []

    def read_excel_headers(self, filename: str) -> list[str]:
        wb = load_workbook(filename=filename, read_only=True, data_only=True)
        try:
            ws = wb.active
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not first_row:
                return []
            return [str(v).strip() for v in first_row if v is not None and str(v).strip()]
        finally:
            wb.close()

    def import_tools(self, filename: str, mapping: dict[str, str]) -> list[dict]:
        wb = load_workbook(filename=filename, read_only=True, data_only=True)
        try:
            ws = wb.active
            raw_headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not raw_headers:
                return []
            headers = [str(v).strip() if v is not None else '' for v in raw_headers]
            header_to_idx = {h: i for i, h in enumerate(headers) if h}

            float_fields = {'geom_x', 'geom_z', 'radius', 'nose_corner_radius', 'drill_nose_angle'}
            int_fields = {'mill_cutting_edges'}
            list_fields = {'support_parts', 'geometry_profiles'}

            imported = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                tool = dict(self.IMPORT_DEFAULTS)

                for field_key, excel_header in mapping.items():
                    if field_key not in self.IMPORT_DEFAULTS:
                        continue
                    idx = header_to_idx.get(excel_header)
                    if idx is None or idx >= len(row):
                        continue
                    raw = row[idx]
                    if raw is None:
                        continue

                    if field_key in float_fields:
                        try:
                            tool[field_key] = float(raw)
                        except Exception:
                            pass
                    elif field_key in int_fields:
                        try:
                            tool[field_key] = int(raw)
                        except Exception:
                            pass
                    elif field_key in list_fields:
                        tool[field_key] = self._parse_json_list_or_empty(raw)
                    else:
                        tool[field_key] = str(raw).strip()

                if not str(tool.get('id', '')).strip():
                    continue

                imported.append(tool)

            return imported
        finally:
            wb.close()

    def export_tools(self, filename: str, tools: list[dict]):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Tools'

        headers = [label for _key, label in self.GENERAL_FIELDS]

        # Header row
        ws.append(headers)
        header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
        header_font = Font(color='FFFFFF', bold=True)
        thin = Side(style='thin', color='D0D7DE')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Stable color mapping for each tool type present in export
        unique_tool_types = []
        for tool in tools or []:
            t = (tool.get('tool_type', '') or '').strip()
            if t and t not in unique_tool_types:
                unique_tool_types.append(t)
        tool_type_fill = {
            t: PatternFill(fill_type='solid', fgColor=self._TOOLTYPE_ROW_COLORS[idx % len(self._TOOLTYPE_ROW_COLORS)])
            for idx, t in enumerate(sorted(unique_tool_types))
        }

        numeric_keys = {'geom_x', 'geom_z', 'radius', 'nose_corner_radius', 'drill_nose_angle'}
        int_keys = {'mill_cutting_edges'}

        # Data rows
        for tool in tools or []:
            row_values = []
            for key, _label in self.GENERAL_FIELDS:
                value = tool.get(key, '')
                if key in numeric_keys:
                    value = self._normalize_number(value)
                elif key in int_keys:
                    try:
                        value = int(value or 0)
                    except Exception:
                        value = value
                row_values.append(value)

            ws.append(row_values)
            row_idx = ws.max_row

            fill = tool_type_fill.get((tool.get('tool_type', '') or '').strip())
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                if fill is not None:
                    cell.fill = fill

                key = self.GENERAL_FIELDS[col_idx - 1][0]
                if key in numeric_keys and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.000'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif key in int_keys and isinstance(cell.value, int):
                    cell.number_format = '0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        # Auto-size columns with sensible min/max limits.
        for col_idx, (_key, _label) in enumerate(self.GENERAL_FIELDS, start=1):
            letter = get_column_letter(col_idx)
            max_len = len(headers[col_idx - 1])
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 48)

        # Slightly taller rows for readability.
        ws.row_dimensions[1].height = 24
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 21

        wb.save(filename)

    def export_tools_to_excel(self, conn, filename):
        # Backward-compatible path for legacy callers.
        rows = conn.execute('SELECT * FROM tools ORDER BY id').fetchall()
        tools = [dict(r) for r in rows]
        self.export_tools(filename, tools)
