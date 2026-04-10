import re
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SERIAL_PATTERN = re.compile(r"^([A-Z]+)(\d{2})(?:/(\d+))?$")


def index_to_letters(index):
    if index < 0:
        raise ValueError("index must be >= 0")
    value = index
    letters = []
    while True:
        value, rem = divmod(value, 26)
        letters.append(chr(ord("A") + rem))
        if value == 0:
            break
        value -= 1
    return "".join(reversed(letters))


class LogbookService:
    _ROW_COLORS = [
        'EAF4FF',
        'EDF8EA',
        'FFF4E8',
        'F5ECFF',
        'EAF8F8',
        'FFF9E3',
        'FDEEEF',
        'ECEFF3',
    ]

    def __init__(self, database):
        self.db = database

    @staticmethod
    def _normalize_like(value):
        return f"%{(value or '').strip()}%"

    def list_entries(self, search="", filters=None, limit=0):
        filters = filters or {}
        like = self._normalize_like(search)
        clauses = [
            "(work_id LIKE ? OR order_number LIKE ? OR batch_serial LIKE ? OR notes LIKE ?)",
        ]
        params = [like, like, like, like]

        work_filter = (filters.get("work_id") or "").strip()
        if work_filter:
            clauses.append("work_id LIKE ?")
            params.append(self._normalize_like(work_filter))

        order_filter = (filters.get("order_number") or "").strip()
        if order_filter:
            clauses.append("order_number LIKE ?")
            params.append(self._normalize_like(order_filter))

        year_filter = (filters.get("year") or "").strip()
        if year_filter:
            clauses.append("substr(date, 1, 4) = ?")
            params.append(year_filter)

        date_from = (filters.get("date_from") or "").strip()
        if date_from:
            clauses.append("date >= ?")
            params.append(date_from)

        date_to = (filters.get("date_to") or "").strip()
        if date_to:
            clauses.append("date <= ?")
            params.append(date_to)

        limit_clause = f" LIMIT {int(limit)}" if int(limit) > 0 else ""
        rows = self.db.conn.execute(
            f"""
            SELECT id, work_id, order_number, quantity, batch_serial, date, notes
            FROM logbook
            WHERE {' AND '.join(clauses)}
            ORDER BY date DESC, id DESC{limit_clause}
            """,
            params,
        ).fetchall()
        return [dict(row) for row in rows]

    def latest_entries_by_work_ids(self, work_ids):
        clean_ids = [str(work_id).strip() for work_id in (work_ids or []) if str(work_id).strip()]
        if not clean_ids:
            return {}

        placeholders = ", ".join("?" for _ in clean_ids)
        rows = self.db.conn.execute(
            f"""
            SELECT l1.id, l1.work_id, l1.order_number, l1.quantity, l1.batch_serial, l1.date, l1.notes
            FROM logbook l1
            JOIN (
                SELECT work_id, MAX(date || '|' || printf('%010d', id)) AS max_key
                FROM logbook
                WHERE work_id IN ({placeholders})
                GROUP BY work_id
            ) latest
              ON latest.work_id = l1.work_id
             AND latest.max_key = l1.date || '|' || printf('%010d', l1.id)
            ORDER BY l1.work_id COLLATE NOCASE ASC
            """,
            clean_ids,
        ).fetchall()
        return {row["work_id"]: dict(row) for row in rows}

    def generate_next_serial(self, work_id, year, quantity=None):
        yy = str(int(year))[-2:]
        rows = self.db.conn.execute(
            "SELECT batch_serial FROM logbook WHERE work_id = ? AND substr(date, 1, 4) = ?",
            (work_id, str(int(year))),
        ).fetchall()
        used_indices = []
        for row in rows:
            serial = (row["batch_serial"] or "").strip().upper()
            match = SERIAL_PATTERN.match(serial)
            if not match:
                continue
            letters, serial_yy, _qty = match.groups()
            if serial_yy != yy:
                continue
            index = 0
            for ch in letters:
                index = index * 26 + (ord(ch) - ord("A") + 1)
            used_indices.append(index - 1)

        next_index = (max(used_indices) + 1) if used_indices else 0
        prefix = index_to_letters(next_index)
        return f"{prefix}{yy}"

    def add_entry(self, work_id, order_number, quantity, notes="", custom_serial="", entry_date=None):
        work_id = (work_id or "").strip()
        if not work_id:
            raise ValueError("work_id is required")
        order_number = (order_number or "").strip()
        qty = int(quantity)
        if qty <= 0:
            raise ValueError("quantity must be > 0")

        use_date = entry_date or date.today().isoformat()
        year = int(use_date[:4])
        serial = (custom_serial or "").strip().upper() or self.generate_next_serial(work_id, year, qty)

        with self.db.conn:
            cur = self.db.conn.execute(
                """
                INSERT INTO logbook (work_id, order_number, quantity, batch_serial, date, notes)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (work_id, order_number, qty, serial, use_date, (notes or "").strip()),
            )
            entry_id = cur.lastrowid
        return self.get_entry(entry_id)

    def get_entry(self, entry_id):
        row = self.db.conn.execute("SELECT * FROM logbook WHERE id = ?", (int(entry_id),)).fetchone()
        return dict(row) if row else None

    def delete_entry(self, entry_id):
        with self.db.conn:
            self.db.conn.execute("DELETE FROM logbook WHERE id = ?", (int(entry_id),))

    def count_entries_for_work(self, work_id: str) -> int:
        row = self.db.conn.execute(
            "SELECT COUNT(*) FROM logbook WHERE work_id = ?", (str(work_id).strip(),)
        ).fetchone()
        return int(row[0]) if row else 0

    def delete_entries_for_work(self, work_id: str):
        with self.db.conn:
            self.db.conn.execute("DELETE FROM logbook WHERE work_id = ?", (str(work_id).strip(),))

    @staticmethod
    def _hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
        hex_color = hex_color.lstrip('#')
        return int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)

    @staticmethod
    def _rgb_to_hex(rgb: tuple[int, int, int]) -> str:
        return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"

    @staticmethod
    def _blend_rgb(rgb1: tuple[int, int, int], rgb2: tuple[int, int, int], ratio: float) -> tuple[int, int, int]:
        ratio = max(0.0, min(1.0, ratio))
        return (
            int(round(rgb1[0] + (rgb2[0] - rgb1[0]) * ratio)),
            int(round(rgb1[1] + (rgb2[1] - rgb1[1]) * ratio)),
            int(round(rgb1[2] + (rgb2[2] - rgb1[2]) * ratio)),
        )

    @staticmethod
    def _days_in_month(year: int, month: int) -> int:
        if month == 12:
            next_month = datetime(year + 1, 1, 1)
        else:
            next_month = datetime(year, month + 1, 1)
        current_month = datetime(year, month, 1)
        return (next_month - current_month).days

    def _date_gradient_fill(self, date_text: str) -> PatternFill:
        # Moderate monthly palette; first day more saturated, last day lighter.
        month_palette = [
            '#5B7BA8', '#7E6FA3', '#6D9A6B', '#5E90A3', '#A28F5F', '#A07A61',
            '#5E96A8', '#7B77A2', '#7D9A66', '#A07966', '#6F86A0', '#8A7A66',
        ]
        try:
            dt = datetime.strptime((date_text or '').strip(), '%Y-%m-%d')
            base_rgb = self._hex_to_rgb(month_palette[dt.month - 1])
            white_rgb = (255, 255, 255)
            dim_rgb = (240, 245, 250)
            month_days = self._days_in_month(dt.year, dt.month)
            progress = (dt.day - 1) / max(1, month_days - 1)
            # Keep saturation moderate by blending base toward dim color first.
            moderated = self._blend_rgb(base_rgb, dim_rgb, 0.35)
            final_rgb = self._blend_rgb(moderated, white_rgb, progress * 0.70)
            return PatternFill(fill_type='solid', fgColor=self._rgb_to_hex(final_rgb))
        except Exception:
            return PatternFill(fill_type='solid', fgColor='EEF3F8')

    @staticmethod
    def _coerce_number(val):
        """Return val as int if it is a pure integer string with no leading zeros, else return as-is."""
        s = str(val).strip() if val is not None else ''
        if s and s.isdigit() and not (len(s) > 1 and s[0] == '0'):
            try:
                return int(s)
            except ValueError:
                pass
        return val

    @staticmethod
    def _format_date_dmy(date_text: str) -> str:
        """Convert ISO date string 'YYYY-MM-DD' to 'DD/MM/YYYY' for display in Excel."""
        try:
            dt = datetime.strptime((date_text or '').strip(), '%Y-%m-%d')
            return dt.strftime('%d/%m/%Y')
        except Exception:
            return date_text or ''

    def export_entries_to_excel(self, entries, output_path, headers=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Logbook"

        headers = list(headers or ["Date", "Serial", "Work ID", "Order", "Quantity", "Notes"])
        ws.append(headers)
        header_fill = PatternFill(fill_type='solid', fgColor='CFE4F8')
        header_font = Font(name='Segoe UI', color='16334E', bold=True)
        thin = Side(style='thin', color='D0D7DE')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        body_font = Font(name='Segoe UI')
        for item in entries:
            ws.append(
                [
                    item.get("work_id", ""),
                    self._coerce_number(item.get("order_number", "")),
                    self._format_date_dmy(item.get("date", "")),
                    item.get("batch_serial", ""),
                    item.get("quantity", 0),
                    item.get("notes", ""),
                ]
            )
            row_idx = ws.max_row
            fill = self._date_gradient_fill(str(item.get('date', '') or '').strip())
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border
                cell.fill = fill
                cell.font = body_font
                if col == 5 and isinstance(cell.value, int):
                    cell.number_format = '0'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        for idx in range(1, len(headers) + 1):
            letter = get_column_letter(idx)
            max_len = len(headers[idx - 1])
            for row_idx in range(2, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=idx).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            notes_col = len(headers)
            col_min = 22 if idx == notes_col else 12
            ws.column_dimensions[letter].width = min(max(max_len + 2, col_min), 48)

        ws.row_dimensions[1].height = 24
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 21

        wb.save(output_path)
