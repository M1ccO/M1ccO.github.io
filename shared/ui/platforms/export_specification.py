"""
Domain-neutral export schema mapper for catalog items (Phase 3 Platform Layer).

This module provides ExportSpecification, a domain-neutral schema mapper for
Excel I/O operations across TOOLS, JAWS, and future domains. Abstracts the
mechanics of field mapping, data type coercion, and import/export without
hardcoding domain-specific logic.

Design:
  - Phase 3: Hardens as shareable abstraction behind adapters
  - Phase 7: Consolidates ToolExportService + JawExportService into single path
  - Preserves backward-compatible field lists and Excel formatting

Usage:
    spec = ExportSpecification(
        domain_name='tools',
        item_service=tool_service,
        fields=[('id', 'Tool ID'), ('description', 'Description'), ...],
        grouping_strategy='by_worksheet',
        group_by_field='tool_head'
    )
    spec.export_to_file('tools.xlsx', items)
    items = spec.import_from_file('tools.xlsx')
"""

from __future__ import annotations

from dataclasses import dataclass, field as dataclass_field
from typing import Any, Callable, Dict, List, Optional
from pathlib import Path

__all__ = [
    'ExportSpecification',
    'ColumnDefinition',
    'ColumnGrouping',
]


@dataclass
class ColumnDefinition:
    """Metadata for a single export column."""

    name: str  # e.g., 'Tool ID'
    field_key: str  # e.g., 'id'
    width: int = 20
    data_type: str = 'text'  # 'text', 'number', 'integer', 'float', 'boolean'
    alignment: str = 'center'  # 'left', 'center', 'right'
    i18n_key: Optional[str] = None  # e.g., 'tool_library.export.header.tool_id'
    numeric_format: Optional[str] = None  # e.g., '0.000' for floats
    hyperlink_target: Optional[str] = None  # e.g., 'holder_link' for 'holder_code'
    hidden: bool = False


@dataclass
class ColumnGrouping:
    """Grouping strategy and grouping metadata."""

    strategy: str = 'none'  # 'none', 'by_field_value', 'by_worksheet'
    group_by_field: Optional[str] = None  # e.g., 'tool_head' for HEAD1/HEAD2 sheets
    group_field_values: List[str] = dataclass_field(default_factory=list)  # e.g. ['HEAD1', 'HEAD2']


class ExportSpecification:
    """
    Domain-neutral export schema mapper for catalog items.

    Responsibilities:
    - Define column schema per domain (field_key → column metadata)
    - Convert item dict ↔ Excel row via coercion functions
    - Group items for export (e.g., by tool_head into separate worksheets)
    - Import items from Excel with field mapping and defaults
    - Preserve backward-compatibility with legacy field lists

    Does NOT depend on Qt or specific Excel formatting; uses openpyxl for I/O only.
    """

    def __init__(
        self,
        domain_name: str,
        item_service: Any,
        fields: List[tuple],
        grouping_strategy: str = 'none',
        group_by_field: Optional[str] = None,
        defaults: Optional[Dict[str, Any]] = None,
        coercers: Optional[Dict[str, Callable]] = None,
        translator: Optional[Callable[[str, str], str]] = None,
    ):
        """
        Initialize export specification.

        Args:
            domain_name: 'tools', 'jaws', etc. Used for i18n key prefixes.
            item_service: Service with list_items(), save_item() methods.
            fields: List of (field_key, display_name) tuples for export columns.
            grouping_strategy: 'none', 'by_worksheet' (use group_by_field).
            group_by_field: Field name to group by (e.g., 'tool_head').
            defaults: Field defaults for import (merged with IMPORT_DEFAULTS).
            coercers: Dict[field_key] → Callable[[value], coerced_value].
            translator: Callable[[i18n_key, default_text], localized_text].
        """
        self.domain_name = domain_name
        self.item_service = item_service
        self.fields = fields
        self.grouping = ColumnGrouping(
            strategy=grouping_strategy,
            group_by_field=group_by_field if grouping_strategy == 'by_worksheet' else None,
        )
        self.defaults = defaults or {}
        self.coercers = coercers or {}
        self._translator = translator or (lambda k, d: d)
        self._column_definitions = self._build_column_definitions()

    def _build_column_definitions(self) -> List[ColumnDefinition]:
        """
        Build column metadata from field list.

        Returns:
            List[ColumnDefinition] with one entry per field in self.fields.
        """
        columns = []
        for field_key, display_name in self.fields:
            i18n_key = f"{self.domain_name}.export.header.{field_key}"
            localized_name = self._translator(i18n_key, display_name)

            col = ColumnDefinition(
                name=localized_name,
                field_key=field_key,
                width=self._infer_width(field_key),
                data_type=self._infer_data_type(field_key),
                numeric_format=self._infer_numeric_format(field_key),
                i18n_key=i18n_key,
            )
            columns.append(col)
        return columns

    def _infer_width(self, field_key: str) -> int:
        """Infer column width from field key (can be overridden per domain)."""
        width_map = {
            'id': 18, 'description': 34, 'tool_type': 24,
            'geom_x': 15, 'geom_z': 15, 'radius': 10,
            'notes': 26, 'holder_code': 24, 'cutting_code': 24,
        }
        return width_map.get(field_key, 20)

    def _infer_data_type(self, field_key: str) -> str:
        """Infer data type from field key."""
        if 'id' in field_key or 'code' in field_key:
            return 'text'
        if any(x in field_key for x in ['geom', 'radius', 'angle', 'diameter', 'length']):
            return 'float'
        if 'edges' in field_key or 'count' in field_key:
            return 'integer'
        return 'text'

    def _infer_numeric_format(self, field_key: str) -> Optional[str]:
        """Infer Excel number format string."""
        if 'angle' in field_key:
            return '0.000'
        if any(x in field_key for x in ['geom', 'radius', 'diameter', 'length']):
            return '0.000'
        if 'edges' in field_key or 'count' in field_key:
            return '0'
        return None

    def get_column_definitions(self) -> List[ColumnDefinition]:
        """
        Get column definitions for export.

        Returns:
            List[ColumnDefinition] with metadata for each export column.
        """
        return self._column_definitions

    def item_to_row(self, item: Dict[str, Any]) -> List[Any]:
        """
        Convert item dict to Excel row values.

        Args:
            item: Dict with field values.

        Returns:
            List[Any] with values in order of self.fields, coerced per column type.
        """
        row = []
        for col in self._column_definitions:
            raw_value = item.get(col.field_key, '')
            coerced = self._coerce_value(raw_value, col.field_key, col.data_type)
            row.append(coerced)
        return row

    def row_to_item(self, row: List[Any]) -> Dict[str, Any]:
        """
        Convert Excel row values back to item dict.

        Args:
            row: List[Any] with values in order of self.fields.

        Returns:
            Dict with field_key → coerced value, plus defaults for missing fields.
        """
        item = dict(self.defaults)
        for idx, col in enumerate(self._column_definitions):
            if idx < len(row):
                raw_value = row[idx]
                coerced = self._coerce_value(raw_value, col.field_key, col.data_type)
                item[col.field_key] = coerced
        return item

    def _coerce_value(self, value: Any, field_key: str, data_type: str) -> Any:
        """
        Coerce value to data type. Uses domain-specific coercers if registered.

        Args:
            value: Raw value from item or Excel cell.
            field_key: Which field is being coerced.
            data_type: Target data type.

        Returns:
            Coerced value, or empty/default on error.
        """
        if field_key in self.coercers:
            return self.coercers[field_key](value)

        if data_type == 'float':
            return self._coerce_float(value, default=0.0)
        elif data_type == 'integer':
            return self._coerce_int(value, default=0)
        elif data_type == 'boolean':
            return bool(value) if value not in (None, '', False, 'False', '0') else False
        else:  # text
            return self._coerce_text(value)

    @staticmethod
    def _coerce_text(value: Any) -> str:
        """Coerce to text, handle None/int/float."""
        if value is None:
            return ''
        if isinstance(value, bool):
            return str(value)
        if isinstance(value, (int, float)):
            return str(value)
        return str(value).strip()

    @staticmethod
    def _coerce_float(value: Any, default: float = 0.0) -> float:
        """Coerce to float, replace ',' with '.'."""
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(',', '.')
        try:
            return float(text) if text else default
        except (ValueError, TypeError):
            return default

    @staticmethod
    def _coerce_int(value: Any, default: int = 0) -> int:
        """Coerce to int."""
        if value is None:
            return default
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        text = str(value).strip().replace(',', '.')
        try:
            return int(float(text)) if text else default
        except (ValueError, TypeError):
            return default

    def export_to_file(self, file_path: str, items: List[Dict[str, Any]]) -> None:
        """
        Export items to Excel file.

        Args:
            file_path: Path to .xlsx file to create/overwrite.
            items: List of item dicts to export.

        Raises:
            ValueError: If items is empty or file_path invalid.
            ImportError: If openpyxl not available (caller responsibility).
        """
        from openpyxl import Workbook

        if not items:
            raise ValueError("Cannot export empty item list")

        wb = Workbook()
        wb.active.title = 'Sheet'

        # For now, single-sheet export. Phase 7 can add grouping logic.
        ws = wb.active
        headers = [col.name for col in self._column_definitions]
        ws.append(headers)

        for item in items:
            row = self.item_to_row(item)
            ws.append(row)

        wb.save(file_path)

    def import_from_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Import items from Excel file.

        Args:
            file_path: Path to .xlsx file to read.

        Returns:
            List[dict] with imported items, coerced per schema.

        Raises:
            FileNotFoundError: If file_path does not exist.
            ImportError: If openpyxl not available (caller responsibility).
        """
        from openpyxl import load_workbook

        if not Path(file_path).exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        imported = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_values = [cell.value for cell in row]
            item = self.row_to_item(row_values)
            if item:
                imported.append(item)

        wb.close()
        return imported
